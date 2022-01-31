import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from pathlib import Path


def rename():
    owd = os.getcwd()
    try:
        plist = os.listdir(r"execute")
        os.chdir(r"execute")
        for pname in plist :
            os.rename(pname, re.sub('[^A-Za-z.]', '' , (pname)))
    finally:
        os.chdir('../')
rename()


def income_statement(trial_balance):
    current_month = '2021-12-31'
    wb = load_workbook('execute/GAAPIS.xlsx')
    ws = wb.active
    dates = pd.Series(['2021-01-31', '2021-02-28', '2021-03-31','2021-04-30','2021-05-31','2021-06-30','2021-07-31','2021-08-31','2021-09-30','2021-10-31','2021-11-30','2021-12-31'])
    idx = pd.Series(['D','E','F','G','H','I','J','K','L','M','N','O'])
    idx.index = dates 
    interest_inc = '10'
    interest_exp = '11'
    gnii = '12'
    pcl = '13'
    nii = '14'
    insurance = '17'
    licensing_income = '18'
    other_income = '19'
    total_other = '20'
    nii_net = '24'
    sb = '27'
    ooe = '28'
    rc = '29'
    toe = '30'
    oibit = '32'
    pfit = '34'
    net_income_id = '36'
    salaries_benefits_geo = 42
    data = pd.read_csv('support/tb-modified.csv', "_")
    data.columns = ["Company","GLAcct", "CurrentMo", "PriorMo", "Abs_Diff"]
    df = data[data.GLAcct.notnull() & data.CurrentMo.notnull() ]
    df['GLAcct'] = df['GLAcct'].str.replace(" ","_")
    df = df.set_index('GLAcct')

    interest_income = round(float(df['Abs_Diff'].loc['Total_Interest_Income'])/1000)*-1
    interest_expense = round(float(df['Abs_Diff'].loc['Total_Interest_Expense'])/1000)
    niibpcl = round(interest_income-interest_expense)
    provision_credit_losses = round(float(df['Abs_Diff'].loc['Total_Provision_for_Credit_Losses'])/1000)
    net_interest_income = (niibpcl-provision_credit_losses)
    insurance_commissions = round(float(df['Abs_Diff'].loc['Total_Insurance_Commissions'])/1000)*-1
    international_income = round(float(df['Abs_Diff'].loc['Total_International_Income'])/1000)
    other_licensing = round(float(df['Abs_Diff'].loc['49801000'])/1000)
    other_LCNG_USB = round(float(df['Abs_Diff'].loc['49802000'])/1000)
    other_LCNG_US_contra = round(float(df['Abs_Diff'].loc['49802100'])/1000)
    other_gift_card = round((float(df['Abs_Diff'].loc[(df['Company'] == '2') & (df.index == '49804000')])+float(df['Abs_Diff'].loc[(df['Company'] == '20') & (df.index == '49804000')]))/1000)
    total_licensing_income = (international_income+other_licensing+other_LCNG_USB+other_LCNG_US_contra+other_gift_card)*-1
    other_lease_income = round(float(df['Abs_Diff'].loc['Total_Lease_Income'])/1000)
    other_investment_income = round(float(df['Abs_Diff'].loc['Total_Investment_Income'])/1000)
    other_billpay_income = round(float(df['Abs_Diff'].loc['49803000'])/1000)
    other_dealershipinsurancefee_income = round(float(df['Abs_Diff'].loc['49811000'])/1000)
    other_misc = round(float(df['Abs_Diff'].loc[(df['Company'] == '2') & (df.index == '49900000')])/1000)
    other_income_v = (other_lease_income+other_investment_income+other_billpay_income+other_dealershipinsurancefee_income+other_misc)*-1
    total_other_income = (insurance_commissions+total_licensing_income+other_income_v)
    net_income = total_other_income+net_interest_income
    salaries_benefits = round(float(df['Abs_Diff'].loc['Total_Salaries_and_Benefits'])/1000)+salaries_benefits_geo
    misc_exp_rec = round((float(df['Abs_Diff'].loc[(df['Company'] == '2') & (df.index == '68991100')])+float(df['Abs_Diff'].loc[(df['Company'] == '13') & (df.index == '68991100')]))/1000)-salaries_benefits_geo
    other_operating_expense = (round(float(df['Abs_Diff'].loc['Total_Operating_Expenses'])/1000)-salaries_benefits-misc_exp_rec)
    total_operating_expenses = salaries_benefits+misc_exp_rec+other_operating_expense
    income_tax_provision = (round(float(df['Abs_Diff'].loc['Total_Provision_for_Income_Taxes'])/1000))

    ws[f'{idx[current_month]}{interest_inc}'].value = interest_income
    ws[f'{idx[current_month]}{interest_exp}'].value = interest_expense
    ws[f'{idx[current_month]}{gnii}'].value = f'={idx[current_month]}{interest_inc}-{idx[current_month]}{interest_exp}'
    ws[f'{idx[current_month]}{pcl}'].value = provision_credit_losses
    ws[f'{idx[current_month]}{nii}'].value = net_interest_income
    ws[f'{idx[current_month]}{insurance}'].value = insurance_commissions
    ws[f'{idx[current_month]}{licensing_income}'].value = total_licensing_income
    ws[f'{idx[current_month]}{other_income}'].value = other_income_v
    ws[f'{idx[current_month]}{total_other}'].value = f'={idx[current_month]}{insurance}+{idx[current_month]}{licensing_income}+{idx[current_month]}{other_income}'
    ws[f'{idx[current_month]}{nii_net}'].value = net_interest_income+total_other_income
    ws[f'{idx[current_month]}{sb}'].value = salaries_benefits
    ws[f'{idx[current_month]}{ooe}'].value = other_operating_expense
    ws[f'{idx[current_month]}{rc}'].value = misc_exp_rec
    ws[f'{idx[current_month]}{toe}'].value = salaries_benefits+other_operating_expense+misc_exp_rec
    ws[f'{idx[current_month]}{oibit}'].value = (net_income-total_operating_expenses)
    ws[f'{idx[current_month]}{pfit}'].value = income_tax_provision
    ws[f'{idx[current_month]}{net_income_id}'].value = (net_income-total_operating_expenses)-income_tax_provision
    wb.save(f'execute/{current_month} GAAP IS.xlsx')